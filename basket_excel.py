import openpyxl

# Define variable to load the dataframe
wb = openpyxl.load_workbook("input_data.xlsx")
import pandas as pd


# Define variable to read sheet
wb1 = wb.active

#Date	HomeTeam	VisitTeam	Spread	Total.y
date_data = []
home_data = []
vizit_data = []
spread_data = []
total_data = []



# Iterate the loop to read the cell values
# i = row j = column
for i in range(3, wb1.max_row,2):
	#for j in range(1, wb1.max_column):
	if int(wb1.cell(row=i,column=1).value) < 1000:
		mmdd='0' +str(wb1.cell(row=i,column=1).value)
		gg= str(int(wb1.cell(row=1, column =1).value[4:8])+1)
	else:
		mmdd = str(wb1.cell(row=i, column=1).value)
		gg = wb1.cell(row=1, column=1).value[4:8]
	date = mmdd[0:2] + "/" + mmdd[2:4]+ "/" +gg
	home = wb1.cell(row=i+1, column=4).value
	vizit = wb1.cell(row=i , column=4).value
	spread =  wb1.cell(row=i + 1, column=11).value
	total =  wb1.cell(row=i, column=11).value
	#print(type(spread), type(total))
	if spread == 'pk':
		pass
	elif total == 'pk':
		spread, total = total, spread
	elif float(spread) < float(total):
		spread = -1*spread

	elif float(spread)> float(total):

		spread, total = total, spread
	else:
		pass
	print(date, i, home, vizit, spread, total)
	date_data.append(date)
	home_data.append(home)
	vizit_data.append(vizit)
	spread_data.append(spread)
	total_data.append(total)

df_basket = pd.DataFrame({'Date': date_data, 'HomeTeam':home_data, 'VisitTeam': vizit_data,	'Spread': spread_data, 'Total.y': total_data})
df_basket.to_csv('basket_data.csv', index=False)